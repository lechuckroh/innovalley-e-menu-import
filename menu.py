import re
from collections import defaultdict
from datetime import datetime, date
from typing import List, Any, Dict, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Menu:
    def __init__(self, day: date, service_time, menu_type, main_title, sides, origins):
        self.day = day
        self.service_time = service_time
        self.menu_type = menu_type
        self.main_title = main_title
        self.sides = sides
        self.origins = origins

    def to_dict(self) -> Dict[str, Any]:
        return {
            "date": format_date(self.day),
            "time": self.service_time,
            "type": self.menu_type,
            "title": self.main_title,
            "sides": self.sides,
            "origins": self.origins,
        }

    def __repr__(self):
        return f"{self.day} [{self.service_time} {self.menu_type}] {self.main_title} : {','.join(self.sides)} {','.join(self.origins)}"


def format_date(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def get_cell_value(ws, col: str, row) -> str:
    return ws[f"{col}{row}"].value


def get_cell_values(ws, col: str, row) -> List[str]:
    value = ws[f"{col}{row}"].value
    values = value.splitlines() if value else []
    return list(map(lambda x: x.strip(), filter(lambda x: bool(x), values)))


def read_menu(ws: Worksheet, day: date, service_time: str, menu_type: str, col: str, row_from: int,
              row_to: int) -> Optional[Menu]:
    """
    (col, row_from) ~ (col, row_to) 영역의 내용을 읽어와서 Menu 객체를 반환한다.
    :param ws: 워크시트
    :param day: 메뉴 날짜
    :param service_time: 조식,중식,석식
    :param menu_type: 메뉴 타입
    :param col: 셀 컬럼
    :param row_from: 셀 시작 행
    :param row_to: 셀 종료 행
    :return: Menu 객체
    """
    lines: List[str] = []
    for row in range(row_from, row_to + 1):
        lines.extend(get_cell_values(ws, col, row))

    # 메뉴 없음
    if not lines:
        return None

    title = lines[0]
    sides: List[str] = []
    origins: List[str] = []
    for line in lines[1:]:
        b, origin = get_origin(line)
        if b:
            origins.append(origin)
        else:
            sides.append(line)

    return Menu(
        day=day,
        service_time=service_time,
        menu_type=menu_type,
        main_title=title,
        sides=sides,
        origins=origins,
    )


def get_origin(s: str) -> (bool, str):
    """
    원산지 문자열인 경우 원산지 문자열을 반환한다.
    """
    regex = r"\((.+:.+산)\)"
    matches = re.search(regex, s)
    if matches:
        return True, matches.groups()[0]
    else:
        return False, ""


def parse_date(s: str) -> date:
    """
    날짜 문자열을 date 객체로 반환한다.
    """
    regex = r"(\d+)\/(\d+)"
    groups = re.search(regex, s).groups()
    if len(groups) == 2:
        now = datetime.now()
        year = now.year
        month = int(groups[0])
        day = int(groups[1])
        if month == 1 and now.month == 12:
            year += 1
        return date(year, month, day)
    raise ValueError(f"failed to parse date: {s}")


def read_day(ws: Worksheet, col: str) -> List[Menu]:
    """
    col 열에 해당하는 날짜의 메뉴들을 읽어온다.
    """
    d: date = parse_date(get_cell_value(ws, col, 3))

    menu_list = list(filter(lambda m: bool(m), [
        read_menu(ws, d, "조식", "A", col, 4, 6),
        read_menu(ws, d, "조식", "TakeOut", col, 7, 7),
        read_menu(ws, d, "중식", "A", col, 8, 13),
        read_menu(ws, d, "중식", "B", col, 14, 18),
        read_menu(ws, d, "중식", "TakeOut1", col, 19, 19),
        read_menu(ws, d, "중식", "TakeOut2", col, 20, 20),
        read_menu(ws, d, "중식", "Plus", col, 21, 21),
        read_menu(ws, d, "석식", "A", col, 22, 27),
        read_menu(ws, d, "석식", "TakeOut1", col, 28, 28),
        read_menu(ws, d, "석식", "TakeOut2", col, 29, 29),
        read_menu(ws, d, "석식", "Plus", col, 30, 30),
    ]))

    # 운영하지 않는 날
    if len(menu_list) <= 1:
        return []

    return menu_list


# 메뉴 엑셀 파일 로드
def load_excel(file: str) -> List[Menu]:
    wb: Workbook = load_workbook(filename=file)
    ws: Worksheet = wb.active

    menus: List[Menu] = []
    menus.extend(read_day(ws, 'D'))
    menus.extend(read_day(ws, 'E'))
    menus.extend(read_day(ws, 'F'))
    menus.extend(read_day(ws, 'G'))
    menus.extend(read_day(ws, 'H'))
    return menus
