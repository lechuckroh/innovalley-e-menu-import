import re
from datetime import datetime, date
from typing import List

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Menu:
    def __init__(self, day: date, service_time, menu_type, main_title, sides, origins):
        self.day = day
        self.service_time = service_time
        self.menu_type = menu_type
        self.main_title = main_title
        self.sides = sides
        self.origins = origins

    def __repr__(self):
        return f"{self.day} [{self.service_time} {self.menu_type}] {self.main_title} : {','.join(self.sides)} {','.join(self.origins)}"


def get_cell_value(ws, col, row) -> str:
    return ws[f"{col}{row}"].value


def get_cell_values(ws, col, row) -> List[str]:
    value = ws[f"{col}{row}"].value
    values = value.splitlines() if value else []
    return list(map(lambda x: x.strip(), filter(lambda x: bool(x), values)))


def read_item(ws: Worksheet, day: date, service_time, menu_type, col, row_from, row_to) -> Menu:
    lines: List[str] = []
    for row in range(row_from, row_to + 1):
        lines.extend(get_cell_values(ws, col, row))

    title = lines[0]
    sides: List[str] = []
    origins: List[str] = []
    for line in lines[1:]:
        b, origin = get_origin(line)
        if b:
            origins.append(origin)
        else:
            sides.append(line)

    return Menu(
        day=day,
        service_time=service_time,
        menu_type=menu_type,
        main_title=title,
        sides=sides,
        origins=origins,
    )


def get_origin(s: str) -> (bool, str):
    regex = r"\((.+:.+산)\)"
    matches = re.search(regex, s)
    if matches:
        return True, matches.groups()[0]
    else:
        return False, ""


def parse_date(s: str) -> date:
    regex = r"(\d+)\/(\d+)"
    groups = re.search(regex, s).groups()
    if len(groups) == 2:
        now = datetime.now()
        year = now.year
        month = int(groups[0])
        day = int(groups[1])
        if month == 1 and now.month == 12:
            year += 1
        return date(year, month, day)
    raise ValueError(f"failed to parse date: {s}")


def read_day(ws: Worksheet, col) -> List[Menu]:
    d: date = parse_date(get_cell_value(ws, col, 3))

    return [
        read_item(ws, d, "조식", "A코너", col, 4, 6),
        read_item(ws, d, "조식", "TakeOut", col, 7, 7),
        read_item(ws, d, "중식", "A코너", col, 8, 13),
        read_item(ws, d, "중식", "B코너", col, 14, 18),
        read_item(ws, d, "중식", "TakeOut1", col, 19, 19),
        read_item(ws, d, "중식", "TakeOut2", col, 20, 20),
        read_item(ws, d, "중식", "Plus", col, 21, 21),
        read_item(ws, d, "석식", "", col, 22, 27),
        read_item(ws, d, "석식", "TakeOut1", col, 28, 28),
        read_item(ws, d, "석식", "TakeOut2", col, 29, 29),
        read_item(ws, d, "석식", "Plus", col, 30, 30),
    ]


def read_all(ws: Worksheet) -> List[Menu]:
    menus: List[Menu] = []
    menus.extend(read_day(ws, 'D'))
    menus.extend(read_day(ws, 'E'))
    menus.extend(read_day(ws, 'F'))
    menus.extend(read_day(ws, 'G'))
    menus.extend(read_day(ws, 'H'))
    return menus


if __name__ == "__main__":
    wb: Workbook = load_workbook(filename='samples/20.03.30.xlsx')
    ws: Worksheet = wb.active
    result = read_all(ws)
    for m in result:
        print(m)
